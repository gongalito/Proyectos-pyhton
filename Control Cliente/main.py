import os
import sys
import glob
import shutil
import locale
import calendar
import pandas as pd
import customtkinter as ctk

from datetime import datetime
from tkinter import scrolledtext
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font



# ------------------------------
# Configuración de rutas
# ------------------------------
BASE_DIR = os.getcwd()  
CONTROL_DIR = os.path.join(BASE_DIR, "CONTROL")
DATA_DIR = os.path.join(BASE_DIR, "Data")
PAGOS_DIR = os.path.join(BASE_DIR, "Pagos")
AGREGAR_PAGO_DIR = os.path.join(BASE_DIR, "Agregar pago")
AGREGAR_CONTROL_DIR = os.path.join(BASE_DIR, "Agregar control")
COPIAS_DIR = os.path.join(DATA_DIR, "Copias de seguridad")
HISTORIAL_PAGOS_DIR = os.path.join(DATA_DIR, "Historial pagos")
HISTORIAL_CONTROL_DIR = os.path.join(DATA_DIR, "Historial control")
AYUDA_DIR = os.path.join(DATA_DIR, "Textos de ayuda")
FACTURASNOPAGADAS_DIR = os.path.join(BASE_DIR, "FacturasNoPagadas")
OUTPUT_FILE_FACTURASNOPAGADAS = os.path.join(BASE_DIR, "FacturasNoPagadas", "FacturasNoPagadas.xlsx")
CONTROL_FILE = os.path.join(CONTROL_DIR, "control.xlsx")
ARCHIVO_CONTROL = os.path.join(CONTROL_DIR, "CONTROL.xlsx")




# Colores
COLOR_PAGADO = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_NO_PAGADO = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")



# Diccionario para traducir meses a español
MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}


# ------------------------------
# Funcion para manejar logs
# ------------------------------
def default_log(msg):
    print(msg)




# ------------------------------
# FUNCIONES DEL PROGRAMA
# ------------------------------
def crear_copia_seguridad(log_callback=None):
    log_callback = log_callback or default_log

    if not os.path.exists(COPIAS_DIR):
        os.makedirs(COPIAS_DIR)

    # Nombre base de la carpeta de backup
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_base = f"CONTROL_{fecha}"
    ruta_backup = os.path.join(COPIAS_DIR, nombre_base)

    # Verificar si ya existe
    contador = 1
    while os.path.exists(ruta_backup):
        nombre_backup = f"{nombre_base} ({contador})"
        ruta_backup = os.path.join(COPIAS_DIR, nombre_backup)
        contador += 1
    else:
        nombre_backup = os.path.basename(ruta_backup)


    # Copiar toda la carpeta CONTROL
    try:
        shutil.copytree("CONTROL", ruta_backup)
        log_callback(f"Copia de seguridad creada: {nombre_backup}\n")
    except Exception as e:
        log_callback(f"[ERROR] No se pudo crear la copia de seguridad: {e}")
        return

    # Limitar a 15 copias 
    copias = [d for d in os.listdir(COPIAS_DIR)
              if os.path.isdir(os.path.join(COPIAS_DIR, d)) and d.startswith("CONTROL_")]
    copias.sort()  # Orden cronológico (por nombre)

    while len(copias) > 15:
        copia_mas_antigua = copias.pop(0)
        ruta_a_eliminar = os.path.join(COPIAS_DIR, copia_mas_antigua)
        try:
            shutil.rmtree(ruta_a_eliminar)
            log_callback(f"Copia antigua eliminada: {copia_mas_antigua}")
        except Exception as e:
            log_callback(f"[ERROR] No se pudo eliminar la copia antigua {copia_mas_antigua}: {e}")




def default_log(msg):
    print(msg)



def ajustar_formato_archivo_pago(ruta_archivo, fecha_pago, log_callback=None):
    log_callback = log_callback or default_log
    df = pd.read_excel(ruta_archivo, dtype=str)
    df.columns = [col.strip().upper() for col in df.columns]

    if df.shape[1] >= 5:
        df.drop(df.columns[4], axis=1, inplace=True)

    # Normalizar columna "NUMERO"
    if "NUMERO" in df.columns:
        def limpiar_numero(x):
            if pd.isna(x):
                return ""
            x = str(x)
            numeros = ''.join(filter(str.isdigit, x))
            return "A" + numeros if numeros else "A"

        df["NUMERO"] = df["NUMERO"].apply(limpiar_numero)


    if "FECHA" in df.columns:
        # Convertir a datetime usando dayfirst=True para dd/mm/yyyy y luego formatear como texto dd/mm/yyyy
        df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
    else:
        raise KeyError(f"No se encontró la columna 'FECHA' en {ruta_archivo}")

    fecha_pago_dt = pd.to_datetime(fecha_pago, errors="coerce")
    fecha_pago_str = fecha_pago_dt.strftime("%d/%m/%Y") if pd.notna(fecha_pago_dt) else fecha_pago
    df["FECHA PAGO"] = fecha_pago_str

    nombre_nuevo = f"pago_{fecha_pago_str.replace('/','-')}.xlsx"
    carpeta = os.path.dirname(ruta_archivo)
    nueva_ruta = os.path.join(carpeta, nombre_nuevo)

    with pd.ExcelWriter(nueva_ruta, engine="openpyxl", date_format="DD/MM/YYYY", datetime_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, index=False)

    if nueva_ruta != ruta_archivo:
        os.remove(ruta_archivo)

    # Ajustar formato de columnas en el Excel
    wb = load_workbook(nueva_ruta)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 8
    wb.save(nueva_ruta)

    return nueva_ruta




def agregar_pago(nuevo_destino, log_callback=None, errores_detallados=None, log_completo=None):
    log_callback = log_callback or default_log
    errores_detallados = errores_detallados or []
    log_completo = log_completo or []

    def log_wrapper(mensaje):
        log_callback(mensaje)
        log_completo.append(mensaje)

    # Fuerzo que las columnas FECHA y FECHA PAGO se lean como texto para evitar errores
    df_pago = pd.read_excel(nuevo_destino, dtype={"FECHA": str, "FECHA PAGO": str})
    df_pago.columns = [col.strip().upper() for col in df_pago.columns]

    for col in ["FECHA", "NUMERO", "MONTO", "PAGADO", "FECHA PAGO"]:
        if col not in df_pago.columns:
            raise KeyError(f"El archivo de pago no tiene la columna {col}")

    color_pagado = PatternFill(start_color="93c47d", end_color="93c47d", fill_type="solid")
    color_diferencia = PatternFill(start_color="f6b26b", end_color="f6b26b", fill_type="solid")

    log_wrapper(f"\n{'-'*60}\nProcesando archivo de pago: {nuevo_destino}\n{'-'*60}\n")

    for _, fila in df_pago.iterrows():
        # Convierto a datetime usando dayfirst=True para asegurar dd/mm/yyyy
        fecha = pd.to_datetime(fila["FECHA"], dayfirst=True, errors="coerce")
        if pd.isna(fecha):
            log_wrapper(f"[ERROR] Fecha inválida en fila: {fila}")
            continue

        anio = fecha.year
        mes_nombre = MESES_ES.get(fecha.month, None)
        if not mes_nombre:
            log_wrapper(f"[ERROR] Mes inválido para fecha {fecha}")
            continue

        numero = fila["NUMERO"]
        monto_base = fila["MONTO"]
        monto_pagado = fila["PAGADO"]
        fecha_pago_val = fila["FECHA PAGO"]

        archivo_control = obtener_ruta_control_por_anio(anio)

        if not os.path.exists(archivo_control):
            descripcion = f"Archivo CONTROL para año {anio} no encontrado"
            errores_detallados.append({
                "archivo": nuevo_destino,
                "tipo": "Pago",
                "numero": numero,
                "fecha": fecha.strftime("%d/%m/%Y"),
                "descripcion": descripcion
            })
            log_wrapper(f"[ERROR] {descripcion}")
            continue

        wb = load_workbook(archivo_control)

        if mes_nombre not in wb.sheetnames:
            descripcion = f"Hoja {mes_nombre} no existe en archivo CONTROL {anio}"
            errores_detallados.append({
                "archivo": nuevo_destino,
                "tipo": "Pago",
                "numero": numero,
                "fecha": fecha.strftime("%d/%m/%Y"),
                "descripcion": descripcion
            })
            log_wrapper(f"[ERROR] {descripcion}")
            continue

        ws = wb[mes_nombre]
        encontrado = False

        for row in ws.iter_rows(min_row=2):
            cell_numero = row[1].value
            if str(cell_numero) == str(numero):
                encontrado = True
                try:
                    monto_base_float = float(monto_base)
                    monto_pagado_float = float(monto_pagado)
                except (ValueError, TypeError):
                    log_wrapper(f"[ERROR] Monto inválido en fila: {fila}")
                    continue

                diferencia = abs(monto_base_float - monto_pagado_float)
                porcentaje_diferencia = diferencia / abs(monto_base_float) if monto_base_float != 0 else 0

                if (
                    monto_base_float == monto_pagado_float or
                    abs(porcentaje_diferencia - 0.22) < 0.01 or
                    abs(porcentaje_diferencia - 0.10) < 0.01
                ):
                    fill_color = color_pagado
                else:
                    fill_color = color_diferencia


                for cell in row:
                    cell.fill = fill_color
                row[3].value = monto_pagado
                row[4].value = fecha_pago_val
                log_wrapper(f"Fila actualizada en CONTROL {anio}: Número {numero}, Fecha {fecha.strftime('%d/%m/%Y')}")
                break

        if not encontrado:
            descripcion = f"Número {numero} con monto {monto_base} no encontrado en hoja {mes_nombre} del CONTROL {anio}"
            errores_detallados.append({
                "archivo": nuevo_destino,
                "tipo": "Pago",
                "numero": numero,
                "fecha": fecha.strftime("%d/%m/%Y"),
                "descripcion": descripcion
            })
            log_wrapper(f"[ERROR] {descripcion}")

        wb.save(archivo_control)




def ajustar_formato_archivo_control(ruta_archivo):
    df = pd.read_excel(ruta_archivo)

    # Elimina columnas
    columnas_a_eliminar = ["Cliente", "Tipodocum", "Documento", "Serie", "Nrdoc", "Cfe_serie", "Nrodoc"]
    for col in columnas_a_eliminar:
        if col in df.columns:
            df.drop(col, axis=1, inplace=True)

    mes, anio = None, None
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
        df["Fecha"] = df["Fecha"].dt.strftime("%d/%m/%Y")
        primera_fecha = pd.to_datetime(df["Fecha"].iloc[0], format="%d/%m/%Y", dayfirst=True)
        mes = int(primera_fecha.month)
        anio = int(primera_fecha.year)

    # Ajusta columna Numero
    if "Cfe_numero" in df.columns:
        df["Cfe_numero"] = df["Cfe_numero"].apply(lambda x: "A" + str(x) if pd.notna(x) else x)
        df.rename(columns={"Cfe_numero": "Numero"}, inplace=True)

    # Ajusta columna Monto
    if "Importe" in df.columns:
        df.rename(columns={"Importe": "Monto"}, inplace=True)

    df["PAGADO"] = 0
    df["FECHA PAGO"] = "-"
    df.columns = [col.upper() for col in df.columns]

    if mes and anio:
        mes_nombre = MESES_ES.get(mes, str(mes).zfill(2))
        nuevo_nombre = f"EDC_{mes_nombre}_{anio}.xlsx"
    else:
        nuevo_nombre = os.path.basename(ruta_archivo)

    carpeta = os.path.dirname(ruta_archivo)
    nueva_ruta = os.path.join(carpeta, nuevo_nombre)

    with pd.ExcelWriter(nueva_ruta, engine="openpyxl", date_format="DD/MM/YYYY", datetime_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, index=False)

    wb = load_workbook(nueva_ruta)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=False)
    wb.save(nueva_ruta)

    if nueva_ruta != ruta_archivo:
        os.remove(ruta_archivo)

    wb = load_workbook(nueva_ruta)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 8
    wb.save(nueva_ruta)

    return nueva_ruta

def agregar_control(archivo_control, archivo_agregar, log_callback=None, errores_detallados=None, log_completo=None):
    log_callback = log_callback or default_log
    errores_detallados = errores_detallados or []
    log_completo = log_completo or []

    def log_wrapper(mensaje):
        log_callback(mensaje)
        log_completo.append(mensaje)

    df_agregar = pd.read_excel(archivo_agregar)
    wb = load_workbook(archivo_control)
    filas_agregadas = 0
    filas_omitidas = 0

    for idx, fila in df_agregar.iterrows():
        fecha = pd.to_datetime(fila["FECHA"], dayfirst=True)
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        mes_nombre = calendar.month_name[fecha.month].upper()

        if mes_nombre not in wb.sheetnames:
            ws = wb.create_sheet(title=mes_nombre)
            ws.append(["FECHA", "NUMERO", "MONTO", "PAGADO", "FECHA PAGO"])
        else:
            ws = wb[mes_nombre]

        numeros_existentes = [str(cell.value) for cell in ws["B"][1:]]
        if str(fila["NUMERO"]) in numeros_existentes:
            filas_omitidas += 1
            errores_detallados.append({
                "archivo": archivo_agregar,
                "tipo": "Control",
                "numero": fila["NUMERO"],
                "fecha": fecha.strftime("%d/%m/%Y"),
                "descripcion": "Número duplicado, fila omitida"
            })
            log_wrapper(f"[ERROR] Archivo: {archivo_agregar}, Número duplicado: {fila['NUMERO']}, Fecha: {fecha.strftime('%d/%m/%Y')}")
            continue

        nueva_fila = [
            fecha.strftime("%d/%m/%Y"),
            fila["NUMERO"],
            fila["MONTO"],
            0,
            "-"
        ]
        ws.append(nueva_fila)
        filas_agregadas += 1
        log_wrapper(f"Fila agregada: Número {fila['NUMERO']}, Fecha {fecha.strftime('%d/%m/%Y')}")

        fill = PatternFill(start_color="FF4040", end_color="FF4040", fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for ws_iter in wb.worksheets:
        for col in ws_iter.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_iter.column_dimensions[col_letter].width = max_length + 8
        for cell in ws_iter[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Elimina la hoja temporal
    if "TEMP" in wb.sheetnames and len(wb.sheetnames) > 1:
        std = wb["TEMP"]
        wb.remove(std)

    wb.save(archivo_control)
    log_wrapper(f"Archivo CONTROL actualizado: {filas_agregadas} filas agregadas, {filas_omitidas} filas omitidas.")



def chequeo_inicio():
    carpetas_verificar_formato = [AGREGAR_PAGO_DIR, AGREGAR_CONTROL_DIR, FACTURASNOPAGADAS_DIR]
    carpetas_verificar_bloqueo = [AGREGAR_PAGO_DIR, AGREGAR_CONTROL_DIR, "CONTROL", FACTURASNOPAGADAS_DIR]

    archivos_incorrectos = []
    archivos_abiertos = []

    # Verificar formatos
    for carpeta in carpetas_verificar_formato:
        for archivo in os.listdir(carpeta):
            ruta_archivo = os.path.join(carpeta, archivo)
            if os.path.isfile(ruta_archivo):
                if not archivo.lower().endswith(".xlsx"):
                    archivos_incorrectos.append(ruta_archivo)

    if archivos_incorrectos:
        mostrar_error_formato(archivos_incorrectos)
        return  # Detener si hay error de formato

    # Verificar archivos abiertos
    for carpeta in carpetas_verificar_bloqueo:
        for carpeta_actual, _, archivos in os.walk(carpeta):
            for archivo in archivos:
                ruta = os.path.join(carpeta_actual, archivo)

                # Ignorar archivos temporales como "~$archivo.xlsx" y no xlsx
                if archivo.startswith("~") or not archivo.lower().endswith(".xlsx"):
                    continue

                try:
                    with open(ruta, "a"):
                        pass  
                except (PermissionError, OSError):
                    archivos_abiertos.append(ruta)

    # Chequear especificamente FacturasNoPagadas.xlsx
    ruta_facturas = os.path.join(FACTURASNOPAGADAS_DIR, "FacturasNoPagadas.xlsx")
    if os.path.exists(ruta_facturas):
        try:
            with open(ruta_facturas, "a"):
                pass
        except (PermissionError, OSError):
            archivos_abiertos.append(ruta_facturas)

    if archivos_abiertos:
        mostrar_error_archivo_abierto(archivos_abiertos)
        return  # Detener si hay archivos en uso

    # Si todo esta bien
    abrir_ventana_iniciar()

def obtener_ruta_control_por_anio(anio):
    carpeta_anio = os.path.join("CONTROL", str(anio))
    os.makedirs(carpeta_anio, exist_ok=True)
    ruta_archivo = os.path.join(carpeta_anio, f"CONTROL_{anio}.xlsx")

    return ruta_archivo


def abrir_control_mas_reciente():
    ruta_busqueda = os.path.join("CONTROL", "*", "CONTROL_*.xlsx")
    archivos_control = glob.glob(ruta_busqueda)

    if not archivos_control:
        print("No se encontró ningún archivo de control para abrir.")
        return

    # Ordenar por fecha de modificacion (ultimo modificado primero)
    archivo_mas_reciente = max(archivos_control, key=os.path.getmtime)

    try:
        os.startfile(archivo_mas_reciente)
    except Exception as e:
        print(f"No se pudo abrir el archivo {archivo_mas_reciente}: {e}")






#----------------------
# FACTUAS NO PAGADAS 
#---------------------


def FacturasNoPagadas():

    color_diferencia = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid")
    color_rojo = PatternFill(start_color="FF4040", end_color="FF4040", fill_type="solid")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "NoPagadas"

    # Encabezado
    encabezado = ["FECHA", "NUMERO", "MONTO", "PAGADO", "FECHAPAGO", "DIFERENCIA", "PORCENTAJE"]
    ws_out.append(encabezado)
    for idx, _ in enumerate(encabezado, start=1):
        ws_out.cell(row=1, column=idx).alignment = Alignment(horizontal="center", vertical="center")

    # Recorrer archivos CONTROL_
    for root, dirs, files in os.walk(CONTROL_DIR):
        for file in files:
            if file.endswith(".xlsx") and file.startswith("CONTROL_"):
                ruta = os.path.join(root, file)
                wb = load_workbook(ruta)
                for ws in wb.worksheets:
                    # Detectar indice de columnas dinamicamente
                    ws_headers = [cell.value.strip().upper() if isinstance(cell.value, str) else "" for cell in ws[1]]
                    try:
                        fecha_idx = ws_headers.index("FECHA")
                        monto_idx = ws_headers.index("MONTO")
                        pagado_idx = ws_headers.index("PAGADO")
                    except ValueError:
                        continue  # Si no encuentra columnas requeridas, pasa a la siguiente hoja

                    for row in ws.iter_rows(min_row=2, values_only=False):
                        fecha_cell = row[fecha_idx]
                        fill_color = fecha_cell.fill.start_color.rgb
                        if fill_color:
                            fill_color = fill_color[-6:].upper()
                        else:
                            continue

                        if fill_color in ["F6B26B", "FF4040"]:
                            # Obtener valores
                            monto = row[monto_idx].value or 0
                            pagado = row[pagado_idx].value or 0
                            diferencia = None
                            porcentaje = None

                            if fill_color == "F6B26B":  
                                diferencia = monto - pagado
                                if monto != 0:
                                    porcentaje_val = (diferencia / monto) * 100
                                    porcentaje = f"%{round(porcentaje_val, 2)}"
                                else:
                                    porcentaje = "%0.00"

                            # Copiar fila
                            nueva_fila = [cell.value for cell in row]
                            nueva_fila.extend([diferencia, porcentaje])
                            ws_out.append(nueva_fila)

                            # Aplicar color y centrar
                            row_out_idx = ws_out.max_row
                            for col_idx in range(1, len(nueva_fila) + 1):
                                cell_out = ws_out.cell(row=row_out_idx, column=col_idx)
                                if fill_color == "F6B26B":
                                    cell_out.fill = color_diferencia
                                else:
                                    cell_out.fill = color_rojo
                                cell_out.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas
    for col in ws_out.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws_out.column_dimensions[col_letter].width = max_length + 8

    os.makedirs(os.path.dirname(OUTPUT_FILE_FACTURASNOPAGADAS), exist_ok=True)
    wb_out.save(OUTPUT_FILE_FACTURASNOPAGADAS)
    os.startfile(OUTPUT_FILE_FACTURASNOPAGADAS)





#-----------------------------------
# ACTUALIZAR CONTROL MANUALMENTE
#------------------------------------
def actualizar_control_manualmente(control_dir):
    
    log_negrita = []

    for root, dirs, files in os.walk(control_dir):
        for file in files:
            if file.endswith(".xlsx") and file.startswith("CONTROL_"):
                ruta_archivo = os.path.join(root, file)
                wb = load_workbook(ruta_archivo)

                for ws in wb.worksheets:
                    headers = [cell.value for cell in ws[1]]

                    for fila_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        fila_info = {headers[i] if i < len(headers) else f"COL{i}": row[i].value
                                     for i in range(len(row))}
                        # Revisar si alguna celda está en negrita
                        fila_en_negrita = any(cell.font and cell.font.bold for cell in row)

                        if fila_en_negrita:
                            log_negrita.append({
                                "archivo": ruta_archivo,
                                "hoja": ws.title,
                                "fila": fila_num,  
                                "datos": fila_info
                            })
    return log_negrita



def marcar_pagadas_desde_log(log_negrita, numero_columna="NUMERO", color_pagado_hex="93c47d"):

    color_pagado = PatternFill(start_color=color_pagado_hex, end_color=color_pagado_hex, fill_type="solid")
    archivos_modificados = set()

    for item in log_negrita:
        ruta_archivo = item["archivo"]
        hoja_nombre = item["hoja"]
        numero_factura = item["datos"].get(numero_columna)

        if numero_factura is None:
            continue  # si no tiene numero, no se puede identificar

        wb = load_workbook(ruta_archivo)
        ws = wb[hoja_nombre]

        # Buscar la fila que coincida con el numero
        headers = [cell.value for cell in ws[1]]
        try:
            numero_idx = headers.index(numero_columna)
        except ValueError:
            continue  

        for row in ws.iter_rows(min_row=2):
            valor = row[numero_idx].value
            if valor == numero_factura:
                # Pintar toda la fila del color de pagado, centrar y quitar negrita
                for cell in row:
                    cell.fill = color_pagado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.font:
                        cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=False,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            color=cell.font.color
                        )
                archivos_modificados.add(ruta_archivo)
                break  # fila encontrada, no buscar más

        wb.save(ruta_archivo)

    return list(archivos_modificados)



def cancelarcontrolmanual():

    for root, dirs, files in os.walk(CONTROL_DIR):
        for file in files:
            if file.endswith(".xlsx") and file.startswith("CONTROL_"):
                ruta = os.path.join(root, file)
                wb = load_workbook(ruta)

                for ws in wb.worksheets:
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if cell.font and cell.font.bold:
                                # Crear nueva fuente igual pero sin negrita
                                nueva_fuente = Font(name=cell.font.name,
                                                    size=cell.font.size,
                                                    bold=False,
                                                    italic=cell.font.italic,
                                                    vertAlign=cell.font.vertAlign,
                                                    underline=cell.font.underline,
                                                    strike=cell.font.strike,
                                                    color=cell.font.color)
                                cell.font = nueva_fuente

                wb.save(ruta)




# --------------------
# FUNCIONES FRONTEND
# --------------------
def abrir_ventana_iniciar():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    ventana = ctk.CTk()
    ventana.title("Inicio")
    ventana.geometry("360x380")  

    # Centrar ventana
    ventana.update_idletasks()
    ancho_ventana = 360
    alto_ventana = 380
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    # Definir funciones de botones
    def ejecutar_main():
        ventana.destroy()
        main()

    def ejecutar_facturas():
        ventana.destroy()
        FacturasNoPagadas()

    def actualizar_manual():
        ventana.destroy()
        pedir_confirmacion_actualizar_manualmente(actualizar_control_manualmente(CONTROL_DIR))

    def ayuda():
        ventana.destroy()
        ventana_ayuda()

    def cerrar():
        ventana.destroy()

    # Crear botones 
    boton_width = 300
    boton_height = 60
    boton_font = ("Arial", 16, "bold")  

    botones = [
        ctk.CTkButton(ventana, text="Control", width=boton_width, height=boton_height, font=boton_font, command=ejecutar_main),
        ctk.CTkButton(ventana, text="Facturas No Pagadas", width=boton_width, height=boton_height, font=boton_font, command=ejecutar_facturas),
        ctk.CTkButton(ventana, text="Actualizar Control Manualmente", width=boton_width, height=boton_height, font=boton_font, command=actualizar_manual),
        ctk.CTkButton(ventana, text="Ayuda", width=boton_width, height=boton_height, font=boton_font, command=ventana_ayuda),
        ctk.CTkButton(ventana, text="Cerrar", width=boton_width, height=boton_height, font=boton_font, command=cerrar)
    ]

    for b in botones:
        b.pack(pady=8) 

    # Indice del boton seleccionado
    seleccion = 0
    botones[seleccion].focus_set()

    def actualizar_borde():
        for i, b in enumerate(botones):
            if i == seleccion:
                b.configure(border_width=2, border_color="white")
            else:
                b.configure(border_width=0)

    actualizar_borde()

    def tecla_presionada(event):
        nonlocal seleccion
        if event.keysym == "Up":
            seleccion = (seleccion - 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Down":
            seleccion = (seleccion + 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Return":
            botones[seleccion].invoke()

    # Permitir seleccionar con el mouse 
    def boton_hover(event, indice):
        nonlocal seleccion
        seleccion = indice
        actualizar_borde()

    for idx, b in enumerate(botones):
        b.bind("<Enter>", lambda e, i=idx: boton_hover(e, i))

    ventana.bind("<Up>", tecla_presionada)
    ventana.bind("<Down>", tecla_presionada)
    ventana.bind("<Return>", tecla_presionada)

    ventana.mainloop()



def pedir_fecha_ventana(nombre_archivo):
    fecha_valida = None

    def validar_fecha(event=None):
        nonlocal fecha_valida
        valor = entry.get()
        try:
            datetime.strptime(valor, "%d/%m/%Y")
            fecha_valida = valor
            ventana.destroy()
        except ValueError:
            entry.configure(fg_color="red")

    def reset_color():
        entry.configure(fg_color=ctk.ThemeManager.theme["CTkEntry"]["fg_color"])

    def formatear_fecha(event):
        if event.keysym in ("BackSpace", "Left", "Right", "Delete"):
            return  

        s = ''.join(filter(str.isdigit, entry.get()))
        nueva = ''
        # Construir la cadena formateada
        if len(s) >= 2:
            nueva += s[:2] + '/'
        else:
            nueva += s

        if len(s) >= 4:
            nueva += s[2:4] + '/'
        elif len(s) > 2:
            nueva += s[2:]

        if len(s) > 4:
            nueva += s[4:8]

        # Evitar sobrescribir si no hay cambios
        if entry.get() != nueva:
            pos = entry.index("insert")
            entry.delete(0, "end")
            entry.insert(0, nueva)
            # Colocar cursor al final
            entry.icursor(len(nueva))

    ventana = ctk.CTk()
    ventana.title("Ingrese fecha de pago")
    ventana.geometry("400x150")
    ventana.resizable(False, False)

    # Centrar ventana
    ventana.update_idletasks()
    ancho_ventana = 400
    alto_ventana = 150
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    max_len = 40
    nombre_mostrar = ("..." + nombre_archivo[-(max_len-3):]) if len(nombre_archivo) > max_len else nombre_archivo

    ctk.CTkLabel(
        ventana,
        text=f"Ingrese la fecha del archivo:\n{nombre_mostrar}",
        wraplength=380,
        justify="center"
    ).pack(pady=(15,5))

    entry = ctk.CTkEntry(ventana, width=200)
    entry.pack(pady=(0,10))

    # Enfocar siempre
    ventana.after(100, lambda: entry.focus())

    # Bind dinamico
    entry.bind("<KeyRelease>", lambda e: [reset_color(), formatear_fecha(e)])
    entry.bind("<Return>", validar_fecha)

    btn = ctk.CTkButton(ventana, text="Aceptar", command=validar_fecha)
    btn.pack(pady=(0,10))

    ventana.protocol("WM_DELETE_WINDOW", lambda: None)
    ventana.mainloop()
    return fecha_valida



def obtener_ruta_disponible(ruta_original: str) -> str:

    if not os.path.exists(ruta_original):
        return ruta_original

    base, ext = os.path.splitext(ruta_original)
    contador = 1
    nueva_ruta = f"{base}({contador}){ext}"
    while os.path.exists(nueva_ruta):
        contador += 1
        nueva_ruta = f"{base}({contador}){ext}"
    return nueva_ruta


def mostrar_errores(errores):
    ventana = ctk.CTk()
    ventana.title("Errores del proceso")
    ventana.geometry("800x400")

    # Centrar ventana en pantalla
    ventana.update_idletasks()
    ancho_ventana = 800
    alto_ventana = 400
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    if errores:
        texto = "\n".join(
            f"[{e['tipo']}] Archivo: {e['archivo']} | Número: {e['numero']} | Fecha: {e['fecha']} | {e['descripcion']}"
            for e in errores
        )
        label = ctk.CTkLabel(ventana, text="Se encontraron errores:", font=("", 16))
        label.pack(pady=10)
        scroll = scrolledtext.ScrolledText(ventana, width=100, height=20)
        scroll.pack(padx=10, pady=10)
        scroll.insert("1.0", texto)
        scroll.configure(state="disabled")
    else:
        label = ctk.CTkLabel(ventana, text="Finalizado exitosamente.", font=("", 16))
        label.pack(pady=60)

    btn = ctk.CTkButton(ventana, text="Cerrar", command=ventana.destroy)
    btn.pack(pady=20)

    ventana.mainloop()

def mostrar_exito_ventana(log_completo=None):
    log_completo = log_completo or []

    # Configuracion inicial
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    ventana = ctk.CTk()
    ventana.title("Proceso finalizado")
    ventana.geometry("450x220")
    ventana.resizable(False, False)

    # Centrar ventana en pantalla
    ventana.update_idletasks()
    ancho_ventana = 450
    alto_ventana = 220
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    label = ctk.CTkLabel(ventana, text="¡Finalizado exitosamente!", font=("Roboto", 20))
    label.pack(pady=(30, 20))

    frame_botones = ctk.CTkFrame(ventana)
    frame_botones.pack(pady=10, fill="x", expand=False)

    # Crear botones
    boton_cerrar = ctk.CTkButton(frame_botones, text="Cerrar", width=120, height=40, command=ventana.destroy)
    boton_detalles = ctk.CTkButton(frame_botones, text="Detalles", width=120, height=40)

    botones = [boton_cerrar, boton_detalles]

    # Funcion para mostrar detalles
    def mostrar_detalles():
        detalles_ventana = ctk.CTkToplevel(ventana)
        detalles_ventana.title("Detalles del log")
        detalles_ventana.geometry("600x400")

        textbox = ctk.CTkTextbox(detalles_ventana, width=580, height=380)
        textbox.pack(padx=10, pady=10, fill="both", expand=True)

        for linea in log_completo:
            textbox.insert("end", linea + "\n")
        textbox.configure(state="disabled")

    boton_detalles.configure(command=mostrar_detalles)

    # Empaquetar botones
    boton_cerrar.pack(side="left", padx=40)
    boton_detalles.pack(side="right", padx=40)

    # Indice del boton seleccionado
    seleccion = 0
    botones[seleccion].focus_set()

    def actualizar_borde():
        for i, b in enumerate(botones):
            if i == seleccion:
                b.configure(border_width=2, border_color="white")
            else:
                b.configure(border_width=0)

    actualizar_borde()

    def tecla_presionada(event):
        nonlocal seleccion
        if event.keysym == "Left":
            seleccion = (seleccion - 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Right":
            seleccion = (seleccion + 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Return":
            botones[seleccion].invoke()

    def boton_hover(event, indice):
        nonlocal seleccion
        seleccion = indice
        actualizar_borde()

    for idx, b in enumerate(botones):
        b.bind("<Enter>", lambda e, i=idx: boton_hover(e, i))

    ventana.bind("<Left>", tecla_presionada)
    ventana.bind("<Right>", tecla_presionada)
    ventana.bind("<Return>", tecla_presionada)

    ventana.mainloop()



def mostrar_error_formato(archivos_incorrectos):
    root = ctk.CTk()
    root.withdraw()  # Oculta la ventana principal temporal

    ventana = ctk.CTkToplevel(root)
    ventana.title("ERROR")
    ventana.geometry("600x350")
    ventana.resizable(False, False)

    # Centrar ventana en pantalla
    ventana.update_idletasks()
    ancho_ventana = 600
    alto_ventana = 350
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    # Etiqueta principal
    etiqueta = ctk.CTkLabel(
        ventana,
        text="Se encontraron archivos con formato incorrecto:",
        wraplength=580,
        justify="center"
    )
    etiqueta.pack(pady=(20, 10), padx=10)

    scroll_frame = ctk.CTkScrollableFrame(ventana, width=580, height=230)
    scroll_frame.pack(padx=10, pady=(0, 10), fill="both", expand=True)

    # Mostrar cada archivo incorrecto
    for archivo in archivos_incorrectos:
        archivo_label = ctk.CTkLabel(
            scroll_frame,
            text=archivo,
            anchor="w",
            wraplength=560,
            justify="left"
        )
        archivo_label.pack(fill="x", padx=5, pady=2)

    # Funcion para cerrar ventana y terminar programa
    def cerrar_programa():
        ventana.destroy()
        sys.exit()

    boton_cerrar = ctk.CTkButton(ventana, text="Cerrar", width=120, height=40, command=cerrar_programa)
    boton_cerrar.pack(pady=(0, 15))

    boton_cerrar.focus_set()
    boton_cerrar.configure(border_width=2, border_color="white")

    def tecla_presionada(event):
        if event.keysym == "Return":
            boton_cerrar.invoke()

    ventana.bind("<Return>", tecla_presionada)

    ventana.grab_set()
    ventana.mainloop()



def mostrar_error_archivo_abierto(archivos_abiertos):
    ctk.set_appearance_mode("dark") 

    root = ctk.CTk()
    root.withdraw()  # Oculta la ventana principal temporal

    ventana = ctk.CTkToplevel(root)
    ventana.title("ERROR")
    ventana.geometry("600x350")
    ventana.resizable(False, False)

    # Centrar ventana en pantalla
    ventana.update_idletasks()
    ancho_ventana = 600
    alto_ventana = 350
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    # Etiqueta principal
    etiqueta = ctk.CTkLabel(
        ventana,
        text="Los siguientes archivos están abiertos y deben cerrarse antes de continuar:",
        wraplength=580,
        justify="center"
    )
    etiqueta.pack(pady=(20, 10), padx=10)

    scroll_frame = ctk.CTkScrollableFrame(ventana, width=580, height=230)
    scroll_frame.pack(padx=10, pady=(0, 10), fill="both", expand=True)

    # Mostrar cada archivo abierto
    for archivo in archivos_abiertos:
        archivo_label = ctk.CTkLabel(
            scroll_frame,
            text=archivo,
            anchor="w",
            wraplength=560,
            justify="left"
        )
        archivo_label.pack(fill="x", padx=5, pady=2)

    # Funcion para cerrar ventana y terminar programa
    def cerrar_programa():
        ventana.destroy()
        sys.exit()  # termina el programa

    boton_cerrar = ctk.CTkButton(ventana, text="Cerrar", width=120, height=40, command=cerrar_programa)
    boton_cerrar.pack(pady=(0, 15))

    boton_cerrar.focus_set()
    boton_cerrar.configure(border_width=2, border_color="white")

    def tecla_presionada(event):
        if event.keysym == "Return":
            boton_cerrar.invoke()

    ventana.bind("<Return>", tecla_presionada)

    ventana.grab_set()
    ventana.mainloop()




def pedir_confirmacion_actualizar_manualmente(log_negrita):
    if not log_negrita:
        ventana_sin_facturas_a_modificar()
        return

    ventana = ctk.CTk()
    ventana.title("Confirmar actualización manual")
    ventana.geometry("800x400")

    # Texto con log
    encabezado = f"{'FECHA':<15} {'NUMERO':<15} {'MONTO':<15} {'PAGADO':<15} {'FECHA PAGO':<15}"
    texto_log = encabezado + "\n" + "="*75 + "\n"

    for fila in log_negrita:
        datos = fila['datos']
        texto_log += f"{str(datos.get('FECHA','')):<15} {str(datos.get('NUMERO','')):<15} " \
                     f"{str(datos.get('MONTO','')):<15} {str(datos.get('PAGADO','')):<15} " \
                     f"{str(datos.get('FECHA PAGO','')):<15}\n"

    scroll = ctk.CTkTextbox(ventana, width=780, height=300)
    scroll.pack(padx=10, pady=10, fill="both", expand=True)
    scroll.insert("1.0", texto_log)
    scroll.configure(state="disabled")  # solo lectura

    frame_botones = ctk.CTkFrame(ventana)
    frame_botones.pack(pady=10)

    # Definir acciones de los botones
    def confirmar():
        marcar_pagadas_desde_log(log_negrita)
        ventana.destroy()
        ventana_control_manual_exito()

    def cancelar():
        cancelarcontrolmanual()
        ventana.destroy()
        ventana_control_manual_cancelar()

    # Crear botones
    botones = [
        ctk.CTkButton(frame_botones, text="Confirmar", width=120, command=confirmar),
        ctk.CTkButton(frame_botones, text="Cancelar", width=120, command=cancelar)
    ]

    botones[0].pack(side="left", padx=20)
    botones[1].pack(side="right", padx=20)

    # Indice del boton seleccionado
    seleccion = 0
    botones[seleccion].focus_set()

    # Funcion para actualizar borde del boton seleccionado
    def actualizar_borde():
        for i, b in enumerate(botones):
            if i == seleccion:
                b.configure(border_width=2, border_color="white")
            else:
                b.configure(border_width=0)

    actualizar_borde()

    # Manejo de teclas
    def tecla_presionada(event):
        nonlocal seleccion
        if event.keysym == "Left":
            seleccion = (seleccion - 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Right":
            seleccion = (seleccion + 1) % len(botones)
            botones[seleccion].focus_set()
            actualizar_borde()
        elif event.keysym == "Return":
            botones[seleccion].invoke()

    def boton_hover(event, indice):
        nonlocal seleccion
        seleccion = indice
        actualizar_borde()

    for idx, b in enumerate(botones):
        b.bind("<Enter>", lambda e, i=idx: boton_hover(e, i))

    ventana.bind("<Left>", tecla_presionada)
    ventana.bind("<Right>", tecla_presionada)
    ventana.bind("<Return>", tecla_presionada)

    ventana.mainloop()





def ventana_sin_facturas_a_modificar():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    ventana = ctk.CTk()
    ventana.title("Sin facturas para modificar")
    ventana.geometry("500x200")
    ventana.resizable(False, False)

    # Centrar ventana
    ventana.update_idletasks()
    ancho_ventana = 500
    alto_ventana = 200
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    # Texto informativo
    label = ctk.CTkLabel(
        ventana,
        text="No hay facturas resaltadas en negrita para marcar como pagadas.",
        wraplength=460,
        justify="center",
        font=("Roboto", 16)
    )
    label.pack(pady=(40, 20))

    def cerrar_programa():
        ventana.destroy()
        sys.exit()  # Finaliza el programa

    boton_cerrar = ctk.CTkButton(ventana, text="Cerrar", width=120, height=40, border_width=2, border_color="white", command=cerrar_programa)
    boton_cerrar.pack(pady=20)


    ventana.bind("<Return>", lambda e: cerrar_programa())

    ventana.mainloop()


def ventana_control_manual_exito():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    ventana = ctk.CTk()
    ventana.title("Control manual")
    ventana.geometry("400x200")
    ventana.resizable(False, False)

    # Centrar ventana
    ventana.update_idletasks()
    ancho_ventana, alto_ventana = 400, 200
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    label = ctk.CTkLabel(ventana, text="Facturas modificadas exitosamente", font=("Roboto", 16))
    label.pack(pady=(50, 20))

    def cerrar_ventana():
        ventana.destroy()

    boton_cerrar = ctk.CTkButton(ventana, text="Cerrar", width=120, height=40,border_width=2, border_color="white",  command=cerrar_ventana)
    boton_cerrar.pack(pady=10)

    ventana.bind("<Return>", lambda e: cerrar_ventana())

    ventana.mainloop()



def ventana_control_manual_cancelar():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    ventana = ctk.CTk()
    ventana.title("Control manual")
    ventana.geometry("400x200")
    ventana.resizable(False, False)

    # Centrar ventana
    ventana.update_idletasks()
    ancho_ventana, alto_ventana = 400, 200
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    label = ctk.CTkLabel(ventana, text="Se quitó el formato a todas las facturas en negrita\nNo se realizaron modificaciones", font=("Roboto", 14), justify="center")
    label.pack(pady=(40, 20))

    boton_cerrar = ctk.CTkButton(ventana, text="Cerrar", width=120, height=40,border_width=2, border_color="white",  command=ventana.destroy)
    boton_cerrar.pack(pady=10)

    ventana.bind("<Return>", lambda e: boton_cerrar.invoke())
    
    ventana.mainloop()



def ventana_ayuda():
    ctk.set_appearance_mode("dark")
    ayuda_vent = ctk.CTk()
    ayuda_vent.title("Ayuda / Guía")

    ayuda_vent.update_idletasks()
    ancho_pantalla = ayuda_vent.winfo_screenwidth()
    alto_pantalla = ayuda_vent.winfo_screenheight()
    ayuda_vent.geometry(f"{ancho_pantalla}x{alto_pantalla}+0+0")

    # Orden deseado de los archivos de ayuda
    orden_archivos = [
        "Informacion.txt",
        "Boton control.txt",
        "Boton facturas no pagadas.txt",
        "Boton agregar control manualmente.txt"
    ]

    texto_completo = ""
    if os.path.exists(AYUDA_DIR):
        for archivo_nombre in orden_archivos:
            ruta = os.path.join(AYUDA_DIR, archivo_nombre)
            if os.path.exists(ruta):
                with open(ruta, "r", encoding="utf-8") as f:
                    contenido = f.read()
                subtitulo = os.path.splitext(archivo_nombre)[0]  # Nombre sin extensión
                texto_completo += f"{subtitulo}\n{'-'*len(subtitulo)}\n{contenido}\n\n"
    else:
        texto_completo = "No se encontró la carpeta de ayuda."

    # Mostrar el texto en un CTkTextbox con scroll
    scroll = ctk.CTkTextbox(ayuda_vent)
    scroll.pack(padx=10, pady=10, fill="both", expand=True)
    scroll.insert("1.0", texto_completo)
    scroll.configure(state="disabled")  # Solo lectura
    # Botón de cerrar
    boton_cerrar = ctk.CTkButton(
        ayuda_vent, text="Cerrar", width=120, height=40,
        border_width=2, border_color="white"
    )
    boton_cerrar.pack(pady=10)
    boton_cerrar.focus_set()  

    def cerrar():
        ayuda_vent.destroy()

    boton_cerrar.configure(command=cerrar)
    ayuda_vent.bind("<Return>", lambda e: boton_cerrar.invoke())

    ayuda_vent.mainloop()




# ------------------------------
# FUNCION PRINCIPAL
# ------------------------------
def main(fecha_pago_dict=None, log_callback=None):
    

    log_completo = []  # Lista para acumular todos los logs
    errores_detallados = []  # Lista para errores detallados

    # Wrapper para que cada log tambien se guarde en log_completo
    def log_wrapper(mensaje):
        if log_callback:
            log_callback(mensaje)
        log_completo.append(mensaje)




    # ---- PROCESAR AGREGAR CONTROL ----
    archivos_control = os.listdir(AGREGAR_CONTROL_DIR)
    if archivos_control:
        for archivo in archivos_control:
            crear_copia_seguridad()
            ruta_archivo = os.path.join(AGREGAR_CONTROL_DIR, archivo)

            # Ajusta formato del archivo de control
            ruta_normalizada = ajustar_formato_archivo_control(ruta_archivo)

            # Extraer anio del archivo ya procesado
            df_temp = pd.read_excel(ruta_normalizada)
            if "FECHA" in df_temp.columns and not df_temp.empty:
                primera_fecha = pd.to_datetime(df_temp["FECHA"].iloc[0], format="%d/%m/%Y", errors="coerce")
                anio = primera_fecha.year
            else:
                anio = "SinAño"

            # Carpeta destino segun anio
            carpeta_anio = os.path.join(HISTORIAL_CONTROL_DIR, str(anio))
            os.makedirs(carpeta_anio, exist_ok=True)

            # Definir ruta destino final
            nombre_destino = os.path.basename(ruta_normalizada)
            nuevo_destino = os.path.join(carpeta_anio, nombre_destino)

            # Si ya existe, buscar un nombre disponible
            nuevo_destino = obtener_ruta_disponible(nuevo_destino)

            # Mover archivo al historial
            shutil.move(ruta_normalizada, nuevo_destino)
            log_wrapper(f"Archivo de control procesado y guardado en historial: {nuevo_destino}")

           # Leer el archivo a agregar
            df_temp = pd.read_excel(nuevo_destino)

            # Agrupar las filas por anio
            df_temp["FECHA"] = pd.to_datetime(df_temp["FECHA"], dayfirst=True, errors="coerce")
            df_por_anio = df_temp.groupby(df_temp["FECHA"].dt.year)

            for anio, df_anio in df_por_anio:
                archivo_control_anio = obtener_ruta_control_por_anio(anio)

                # Crear archivo si no existe
                if not os.path.exists(archivo_control_anio):
                    wb_nuevo = Workbook()
                    hoja = wb_nuevo.active
                    hoja.title = "TEMP"
                    encabezados = ["FECHA", "NUMERO", "MONTO", "PAGADO", "FECHA PAGO"]
                    hoja.append(encabezados)
                    wb_nuevo.save(archivo_control_anio)
                    log_wrapper(f"Archivo CONTROL creado: {archivo_control_anio}")

                # Guardar df_anio temporal a un archivo para usar en agregar_control()
                ruta_temp = f"temp_{anio}.xlsx"
                df_anio.to_excel(ruta_temp, index=False)

                agregar_control(archivo_control_anio, ruta_temp, log_wrapper, errores_detallados)

                # Eliminar temporal
                os.remove(ruta_temp)

    else:
        log_wrapper("No hay archivos en 'Agregar control' para procesar.\n")



    # ---- PROCESAR AGREGAR PAGOS ----
    archivos_pagos = os.listdir(AGREGAR_PAGO_DIR)
    if archivos_pagos:
        for archivo in archivos_pagos:
            crear_copia_seguridad(log_wrapper)
            ruta_archivo = os.path.join(AGREGAR_PAGO_DIR, archivo)

            # Obtener fecha del pago
            if fecha_pago_dict and archivo in fecha_pago_dict:
                fecha_pago_input = fecha_pago_dict[archivo]
            else:
                fecha_pago_input = pedir_fecha_ventana(archivo)  

            # Convertir a formato YYYY-MM-DD para internal
            fecha_pago = datetime.strptime(fecha_pago_input, "%d/%m/%Y").strftime("%Y-%m-%d")

            # Ajustar formato del archivo de pago
            ruta_normalizada = ajustar_formato_archivo_pago(ruta_archivo, fecha_pago, log_wrapper)

            # Extraer anio y mes
            fecha_dt = datetime.strptime(fecha_pago, "%Y-%m-%d")
            anio = str(fecha_dt.year)
            mes_nombre = MESES_ES[fecha_dt.month]

            # Crear subcarpetas por anio y mes en HISTORIAL_PAGOS
            carpeta_anio = os.path.join(HISTORIAL_PAGOS_DIR, anio)
            carpeta_mes = os.path.join(carpeta_anio, mes_nombre)
            os.makedirs(carpeta_mes, exist_ok=True)

            # Obtener ruta disponible para no sobrescribir archivos existentes
            nuevo_destino = os.path.join(carpeta_mes, os.path.basename(ruta_normalizada))
            nuevo_destino = obtener_ruta_disponible(nuevo_destino)

            # Mover archivo al historial
            os.rename(ruta_normalizada, nuevo_destino)
            log_wrapper(f"Archivo de pago procesado y guardado en historial: {nuevo_destino}\n")

            # Agregar al archivo CONTROL principal
            agregar_pago(nuevo_destino, log_wrapper, errores_detallados)
    else:
        log_wrapper("No hay archivos en 'Agregar pagos' para procesar.\n")

    log_wrapper("Proceso finalizado.")


    # ---- MOSTRAR RESULTADO ----
    if errores_detallados:
        mostrar_errores_ventana(errores_detallados, log_completo)
    else:
        mostrar_exito_ventana(log_completo)
    
    abrir_control_mas_reciente()



# Ejecutar directamente si se llama desde consola
if __name__ == "__main__":
    chequeo_inicio()

